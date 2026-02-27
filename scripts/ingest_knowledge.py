import argparse
import html
import json
import os
import re
import urllib.request
import zipfile
from pathlib import Path


try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional dependency
    PdfReader = None

try:
    import openpyxl
except Exception:  # pragma: no cover - optional dependency
    openpyxl = None

try:
    import xlrd
except Exception:  # pragma: no cover - optional dependency
    xlrd = None

try:
    from pdf2image import convert_from_path
    import pytesseract
except Exception:  # pragma: no cover - optional dependency
    convert_from_path = None
    pytesseract = None


def normalize_text(text: str) -> str:
    text = html.unescape(text)
    text = re.sub(r"\r", "", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = re.sub(r"[ \t]{2,}", " ", text)
    return text.strip()


def docx_to_text(path: Path) -> str:
    with zipfile.ZipFile(path) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", errors="ignore")
    xml = re.sub(r"</w:p>", "\n", xml)
    xml = re.sub(r"<[^>]+>", "", xml)
    return normalize_text(xml)


def pdf_to_text(path: Path) -> str:
    if PdfReader is None:
        raise RuntimeError("Missing dependency: pypdf")

    reader = PdfReader(str(path))
    chunks = []
    for page in reader.pages:
        chunks.append(page.extract_text() or "")
    text = normalize_text("\n".join(chunks))

    if len(text) >= 200:
        return text

    # OCR fallback for scanned PDFs when optional dependencies are available.
    if convert_from_path is not None and pytesseract is not None:
        images = convert_from_path(str(path), dpi=200)
        ocr_chunks = [pytesseract.image_to_string(image, lang="por") for image in images]
        ocr_text = normalize_text("\n".join(ocr_chunks))
        if len(ocr_text) > len(text):
            return ocr_text

    return text


def xlsx_to_text(path: Path) -> str:
    if openpyxl is None:
        raise RuntimeError("Missing dependency: openpyxl")

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    lines = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        lines.append(f"# Aba: {sheet_name}")
        for row in ws.iter_rows(values_only=True):
            values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if values:
                lines.append(" | ".join(values))
    return normalize_text("\n".join(lines))


def xls_to_text(path: Path) -> str:
    if xlrd is None:
        raise RuntimeError("Missing dependency: xlrd")

    wb = xlrd.open_workbook(path)
    lines = []
    for sheet in wb.sheets():
        lines.append(f"# Aba: {sheet.name}")
        for r in range(sheet.nrows):
            values = []
            for c in range(sheet.ncols):
                value = sheet.cell_value(r, c)
                if value is None:
                    continue
                text = str(value).strip()
                if text:
                    values.append(text)
            if values:
                lines.append(" | ".join(values))
    return normalize_text("\n".join(lines))


def classify_domain(file_name: str) -> str:
    lower = file_name.lower()
    if "swot" in lower or "fofa" in lower:
        return "swot_modelos"
    if "checklist" in lower:
        return "checklists_operacionais"
    if "atendimento" in lower or "aida" in lower:
        return "padrao_atendimento"
    if "promoc" in lower or "podcast" in lower or "roteiro" in lower:
        return "promocoes_marketing"
    if "nr" in lower or "anp" in lower or "inmetro" in lower or "procon" in lower:
        return "compliance_referencias"
    return "geral"


def classify_tags(file_name: str) -> list[str]:
    lower = file_name.lower()
    tags = []
    mapping = {
        "swot": ["swot", "estrategia"],
        "checklist": ["checklist", "operacao"],
        "atendimento": ["atendimento", "vendas"],
        "podcast": ["conteudo", "treinamento"],
        "promoc": ["promocao", "marketing"],
        "nr": ["compliance"],
        "anp": ["compliance"],
        "inmetro": ["compliance"],
        "procon": ["compliance"],
    }
    for key, values in mapping.items():
        if key in lower:
            tags.extend(values)
    return sorted(set(tags))


def extract_text(path: Path) -> tuple[str, str]:
    ext = path.suffix.lower()
    if ext == ".docx":
        return docx_to_text(path), "docx"
    if ext == ".pdf":
        return pdf_to_text(path), "pdf"
    if ext == ".xlsx":
        return xlsx_to_text(path), "xlsx"
    if ext == ".xls":
        return xls_to_text(path), "xls"
    raise RuntimeError(f"Unsupported file extension: {ext}")


def upload_document(api_base: str, admin_token: str, title: str, source: str, text: str, metadata: dict) -> None:
    url = f"{api_base.rstrip('/')}/api/admin/knowledge/upload"
    payload = json.dumps(
        {
            "title": title,
            "source": source,
            "text": text,
            "version": "v2",
            "domain": metadata.get("domain"),
            "tags": metadata.get("tags", []),
            "priority": metadata.get("priority", 3),
            "metadata": metadata,
        }
    ).encode("utf-8")

    request = urllib.request.Request(
        url,
        data=payload,
        method="POST",
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {admin_token}",
        },
    )

    with urllib.request.urlopen(request) as response:
        body = response.read().decode("utf-8")
        print(f"Uploaded {title}: {response.status} {body}")


def should_ingest(path: Path) -> bool:
    return path.suffix.lower() in {".docx", ".pdf", ".xlsx", ".xls"} and not path.name.startswith("~$")


def main() -> None:
    parser = argparse.ArgumentParser(description="Ingest DOCX/PDF/XLSX/XLS files into Rocha Turbo knowledge base.")
    parser.add_argument(
        "--folder",
        default=r"C:\Users\pcram\OneDrive\Documentos\Clientes\Turbo Rocha\Base de Dados",
        help="Folder with files to ingest.",
    )
    parser.add_argument("--api-base", default=os.environ.get("APP_BASE_URL", "http://localhost:3000"))
    parser.add_argument("--admin-token", default=os.environ.get("ADMIN_API_TOKEN", ""))
    parser.add_argument("--min-length", type=int, default=20)
    args = parser.parse_args()

    if not args.admin_token:
        raise SystemExit("ADMIN_API_TOKEN is required (argument or environment variable).")

    folder = Path(args.folder)
    if not folder.exists():
        raise SystemExit(f"Folder not found: {folder}")

    files = [file for file in folder.glob("**/*") if file.is_file() and should_ingest(file)]
    if not files:
        print("No supported files found.")
        return

    for file in sorted(files):
        try:
            text, parser_type = extract_text(file)
        except Exception as error:
            print(f"Skipping {file.name}: extraction error ({error})")
            continue

        if len(text) < args.min_length:
            print(f"Skipping {file.name}: not enough text extracted.")
            continue

        metadata = {
            "domain": classify_domain(file.name),
            "tags": classify_tags(file.name),
            "priority": 3,
            "parser": parser_type,
            "extension": file.suffix.lower(),
            "filename": file.name,
            "size_bytes": file.stat().st_size,
        }

        try:
            upload_document(
                api_base=args.api_base,
                admin_token=args.admin_token,
                title=file.stem,
                source=str(file),
                text=text,
                metadata=metadata,
            )
        except Exception as error:
            print(f"Failed upload {file.name}: {error}")


if __name__ == "__main__":
    main()
